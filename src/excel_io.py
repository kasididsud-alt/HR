import io
import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill


WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")


def read_excel_bytes(uploaded_file) -> pd.DataFrame:
    return pd.read_excel(uploaded_file)


def _apply_basic_sheet_format(ws, df: pd.DataFrame) -> None:
    if df is None:
        return

    ws.freeze_panes = "A2"
    if len(df.columns):
        ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, col_name in enumerate(df.columns, start=1):
        values = [str(col_name)]
        if not df.empty:
            values.extend(df[col_name].dropna().astype(str).head(200).tolist())
        width = min(max(max((len(v) for v in values), default=8) + 2, 10), 42)
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width


def _apply_highlights(
    ws,
    df: pd.DataFrame,
    specs: list[dict] | None,
) -> None:
    if df is None or not specs:
        return

    col_map = {str(col): pos for pos, col in enumerate(df.columns, start=1)}
    max_col = max(len(df.columns), 1)

    for spec in specs:
        row_pos = spec.get("row")
        if row_pos is None:
            continue

        excel_row = int(row_pos) + 2
        fill_name = spec.get("fill", "warn")
        fill = ERROR_FILL if fill_name == "error" else WARN_FILL
        comment_text = str(spec.get("comment", "") or "").strip()

        if spec.get("entire_row"):
            target_cols = range(1, max_col + 1)
        else:
            target_cols = [
                col_map[col]
                for col in spec.get("columns", [])
                if col in col_map
            ]

        for col_idx in target_cols:
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.fill = fill
            if comment_text and cell.comment is None:
                cell.comment = Comment(comment_text, "SalaryCalc")


def _highlight_required_blanks(ws, df: pd.DataFrame, columns: list[str]) -> None:
    if df is None or df.empty:
        return

    col_map = {str(col): pos for pos, col in enumerate(df.columns, start=1)}
    for col_name in columns:
        col_idx = col_map.get(col_name)
        if not col_idx:
            continue
        for row_pos, value in enumerate(df[col_name].tolist(), start=2):
            text = "" if pd.isna(value) else str(value).strip()
            if text:
                continue
            cell = ws.cell(row=row_pos, column=col_idx)
            cell.fill = REQUIRED_FILL
            if cell.comment is None:
                cell.comment = Comment("กรอกข้อมูลช่องนี้ก่อนนำไปใช้", "SalaryCalc")


def to_excel_bytes(
    sheets: dict[str, pd.DataFrame],
    highlights: dict[str, list[dict]] | None = None,
    required_blank_columns: dict[str, list[str]] | None = None,
) -> bytes:
    buf = io.BytesIO()
    highlights = highlights or {}
    required_blank_columns = required_blank_columns or {}

    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        for name, df in sheets.items():
            sheet_name = name[:31]
            df.to_excel(w, index=False, sheet_name=sheet_name)
            ws = w.sheets[sheet_name]
            _apply_basic_sheet_format(ws, df)
            _apply_highlights(ws, df, highlights.get(name) or highlights.get(sheet_name))
            _highlight_required_blanks(
                ws,
                df,
                required_blank_columns.get(name)
                or required_blank_columns.get(sheet_name)
                or [],
            )
    return buf.getvalue()
